# models/settings.py

import pandas as pd
from typing import List, Dict, Any, Optional
from utils.logger_config import logger
from models.soft_account import SoftAccount
import re
import json
import os
from models.status_enum import AccountStatus
from eth_account import Account
import secrets
from openpyxl import load_workbook
import numpy as np
from web3 import Web3


class UserSettingsParser:
    """
    A class to parse user settings from an Excel file with "Main" and "Lombard" sheets.
    """

    def __init__(self, file_path: str = './Soft_settings.xlsx'):
        """
        Initializes the UserSettingsParser.

        Args:
            file_path (str): The path to the Excel file.
        """
        self.file_path = file_path
        self.accounts = []  # List to store SoftAccount instances
        logger.info("Initializing UserSettingsParser")
        self.load_settings()
        self.load_status()

    def load_settings(self):
        """
        Loads and parses the settings from the Excel file.
        """
        try:
            logger.info(f"Loading settings from {self.file_path}")
            # Read the Excel file
            main_df = pd.read_excel(self.file_path, sheet_name='Main')
            lombard_df = pd.read_excel(self.file_path, sheet_name='Lombard')

            # Ensure that both sheets have the same number of rows
            if len(main_df) != len(lombard_df):
                raise ValueError("The 'Main' and 'Lombard' sheets must have the same number of rows.")

            # Iterate over each row (account)
            for index in main_df.index:
                main_row = main_df.loc[index]
                lombard_row = lombard_df.loc[index]

            # Parse and validate the account settings
                account_settings = self.parse_account_settings(main_row, lombard_row, index + 2, main_df, index)  # +2 for Excel row number
                soft_account = SoftAccount(account_settings)
                self.accounts.append(soft_account)

        except Exception as e:
            logger.error(f"Error loading settings: {e}")
            raise Exception(f"Error loading settings: {e}")

    def parse_account_settings(self, main_row: pd.Series, lombard_row: pd.Series, row_number: int, main_df: pd.DataFrame, df_index: int) -> Dict[str, Any]:
        """
        Parses and validates the settings for a single account.

        Args:
            main_row (pd.Series): The row from the 'Main' sheet.
            lombard_row (pd.Series): The row from the 'Lombard' sheet.
            row_number (int): The Excel row number (used for error messages).

        Returns:
            Dict[str, Any]: A dictionary containing the parsed and validated settings.
        """
        account = {}
        logger.debug(f"Parsing account settings for row {row_number}")

        # Parse 'Main' sheet
        try:
            # 1.a. 'private_key' (required)
            private_key = main_row.get('private_key')
            if pd.isna(private_key):
                raise ValueError(f"Row {row_number}: 'private_key' is required.")
            if not isinstance(private_key, str):
                raise ValueError(f"Row {row_number}: 'private_key' must be a string.")
            if private_key.strip().lower() == 'gen':
                logger.info(f"Row {row_number}: Generating new EVM wallet.")
                # Generate new EVM wallet
                new_account = Account.create(secrets.token_hex(32))
                generated_private_key = '0x' + str(new_account.key.hex())
                account['private_key'] = generated_private_key
                # Write the generated private key back to the Excel file
                self.update_private_key_in_excel(main_df, df_index, generated_private_key)
                logger.info(f"Generated new EVM wallet for row {row_number}.")
            else:
                account['private_key'] = private_key

            # **NEW**: Parse 'proxy' (optional)
            # 1.b. 'proxy'
            proxy = main_row.get('proxy')
            if pd.isna(proxy) or proxy == '':
                account['proxy'] = None
            else:
                if not isinstance(proxy, str):
                    raise ValueError(f"Row {row_number}: 'proxy' must be a string or empty.")
                # Proxy format validation
                proxy_pattern = r'^\w+:\w+@\d+\.\d+\.\d+\.\d+:\d+$'
                if not re.match(proxy_pattern, proxy):
                    raise ValueError(f"Row {row_number}: 'proxy' must be in the format 'login:password@ip:port'")
                account['proxy'] = proxy

            # 1.c. 'max_gas_gwei' (optional)
            max_gas_gwei = main_row.get('max_gas_gwei')
            if pd.isna(max_gas_gwei):
                account['max_gas_gwei'] = None  # No limitation
            else:
                if max_gas_gwei is not None and not isinstance(max_gas_gwei, (np.integer)):
                    raise ValueError(f"Row {row_number}: 'max_gas_gwei' must be an integer or empty. The current type is {type(max_gas_gwei)}")
                account['max_gas_gwei'] = max_gas_gwei

            # 1.d. 'exchange' (required)
            exchange = main_row.get('exchange')
            if pd.isna(exchange):
                raise ValueError(f"Row {row_number}: 'exchange' is required.")
            if exchange not in ['OKX', 'Bitget']:
                raise ValueError(f"Row {row_number}: 'exchange' must be 'OKX' or 'Bitget'.")
            account['exchange'] = exchange

            # 1.e. 'exchange_api_key' (required)
            exchange_api_key = main_row.get('exchange_api_key')
            if pd.isna(exchange_api_key):
                raise ValueError(f"Row {row_number}: 'exchange_api_key' is required.")
            account['exchange_api_key'] = str(exchange_api_key)

            # 1.f. 'exchange_secret_key' (required)
            exchange_secret_key = main_row.get('exchange_secret_key')
            if pd.isna(exchange_secret_key):
                raise ValueError(f"Row {row_number}: 'exchange_secret_key' is required.")
            account['exchange_secret_key'] = str(exchange_secret_key)

            # 1.g. 'exchange_passphrase' (required)
            exchange_passphrase = main_row.get('exchange_passphrase')
            if pd.isna(exchange_passphrase):
                raise ValueError(f"Row {row_number}: 'exchange_passphrase' is required.")
            account['exchange_passphrase'] = str(exchange_passphrase)

        except ValueError as ve:
            logger.error(f"Error parsing 'Main' sheet: {ve}")
            raise ve

        # Parse 'Lombard' sheet
        try:
            # 2.a. 'generate_btc_address' (required)
            generate_btc_address = lombard_row.get('generate_btc_address')
            if pd.isna(generate_btc_address):
                raise ValueError(f"Row {row_number}: 'generate_btc_address' is required.")
            if generate_btc_address not in [0, 1]:
                raise ValueError(f"Row {row_number}: 'generate_btc_address' must be 0 or 1.")
            account['generate_btc_address'] = int(generate_btc_address) if generate_btc_address is not None else 0

            # 2.b. 'btc_address'
            btc_address = lombard_row.get('btc_address')
            if account['generate_btc_address'] == 1:
                if not pd.isna(btc_address):
                    # Our software will fill this cell later
                    account['btc_address'] = None
                else:
                    account['btc_address'] = None
            else:
                if pd.isna(btc_address):
                    raise ValueError(f"Row {row_number}: 'btc_address' is required when 'generate_btc_address' is 0.")
                account['btc_address'] = str(btc_address)

            # 2.c. 'min_BTC' (required)
            min_BTC = lombard_row.get('min_BTC')
            if pd.isna(min_BTC):
                raise ValueError(f"Row {row_number}: 'min_BTC' is required.")
            if not isinstance(min_BTC, (int, float)):
                raise ValueError(f"Row {row_number}: 'min_BTC' must be a float.")
            if min_BTC < 0.0002:
                raise ValueError(f"Row {row_number}: 'min_BTC' must be greater than 0.0002.")
            account['min_BTC'] = float(min_BTC)

            # 2.d. 'max_BTC' (required)
            max_BTC = lombard_row.get('max_BTC')
            if pd.isna(max_BTC):
                raise ValueError(f"Row {row_number}: 'max_BTC' is required.")
            if not isinstance(max_BTC, (int, float)):
                raise ValueError(f"Row {row_number}: 'max_BTC' must be a float.")
            if max_BTC < account['min_BTC']:
                raise ValueError(f"Row {row_number}: 'max_BTC' must be greater than or equal to 'min_BTC'.")
            account['max_BTC'] = float(max_BTC)

            # 2.e. 'restaking_LBTC' (required)
            restaking_LBTC = lombard_row.get('restaking_LBTC')
            if pd.isna(restaking_LBTC):
                raise ValueError(f"Row {row_number}: 'restaking_LBTC' is required.")
            if restaking_LBTC is None:
                raise ValueError(f"Row {row_number}: 'restaking_LBTC' cannot be None.")
            account['restaking_LBTC'] = int(restaking_LBTC)

            # 2.f. 'Defi_Vault'
            defi_vault = lombard_row.get('Defi_Vault')
            # 2.g. 'Etherfi'
            etherfi = lombard_row.get('Etherfi')
            # 2.h. 'Pendle'
            pendle = lombard_row.get('Pendle')

            # Validate vault selection
            vaults = {
                'Defi_Vault': defi_vault,
                'Etherfi': etherfi,
                'Pendle': pendle
            }

            if account['restaking_LBTC'] == 1:
                selected_vaults = [k for k, v in vaults.items() if v == 1]
                if len(selected_vaults) != 1:
                    raise ValueError(f"Row {row_number}: Exactly one vault must be selected when 'restaking_LBTC' is 1.")
                account['selected_vault'] = selected_vaults[0]
            else:
                # Ensure vault selections are empty or zero
                for vault_name, vault_value in vaults.items():
                    if not pd.isna(vault_value) and vault_value != 0:
                        raise ValueError(f"Row {row_number}: '{vault_name}' must be empty or 0 when 'restaking_LBTC' is 0.")
                account['selected_vault'] = None

        except ValueError as ve:
            logger.error(f"Error parsing 'Lombard' sheet: {ve}")
            raise ve

        logger.debug(f"Parsed account settings for row {row_number}")
        return account
    
    def update_private_key_in_excel(self, main_df: pd.DataFrame, df_index: int, private_key: str):
        """
        Updates the private key in the 'Soft_settings.xlsx' file for the given account.

        Args:
            main_df (pd.DataFrame): The DataFrame for the 'Main' sheet.
            df_index (int): The index of the current row in the DataFrame.
        private_key (str): The generated private key to write back.
    """
        logger.info("Updating private key in Soft_settings.xlsx")
        settings_file = self.file_path
        try:
            # Load the workbook
            wb = load_workbook(settings_file)
            ws = wb['Main']

            # Find the column index for 'private_key'
            header = [cell.value for cell in ws[1]]  # Get header row values
            try:
                private_key_col_idx = header.index('private_key') + 1  # +1 because openpyxl is 1-indexed
            except ValueError:
                raise Exception("Column 'private_key' not found in 'Main' sheet.")

            # Update the cell value
            # df_index corresponds to the DataFrame index, which starts from 0
            # Excel rows start from 1, with the header at row 1
            excel_row = df_index + 2  # +2 accounts for header row and zero-based index

            ws.cell(row=excel_row, column=private_key_col_idx, value=private_key)

            # Save the workbook
            wb.save(settings_file)
            logger.info("Private key updated in Soft_settings.xlsx")
        except Exception as e:
            logger.error(f"Error updating private key in Soft_settings.xlsx: {e}")
            raise

    def get_accounts(self) -> List[SoftAccount]:
        """
        Returns the list of parsed account settings.

        Returns:
            List[SoftAccount]: A list of SoftAccount instances.
        """
        return self.accounts

    def load_status(self, status_file: str = 'status.json'):
        """
        Loads the account statuses from a JSON file.
        
        Args:
            status_file (str): The path to the status file.
        """
        if os.path.exists(status_file):
            logger.info(f"Loading account statuses from {status_file}")
            with open(status_file, 'r') as f:
                status_data = json.load(f)
            # Expecting status_data to be a dictionary mapping private_key to status info
            for account in self.accounts:
                private_key = account.settings.get('private_key')
                if private_key and private_key in status_data:
                    data = status_data[private_key]
                    account_status = data.get('status')
                    if account_status:
                        account.status = AccountStatus(account_status)
                    account.btc_address = data.get('btc_address')
                    account.withdrawal_id = data.get('withdrawal_id')
                    account.transaction_hash = data.get('transaction_hash')
                    logger.debug(f"Loaded status for account {Web3.eth.account.from_key(private_key).address}: {account.status}")
                else:
                    logger.info(f"No existing status for account {Web3.eth.account.from_key(private_key).address}. Setting to INIT.")
                    account.status = AccountStatus.INIT
        else:
            logger.info(f"No existing status file found at {status_file}")

    def save_status(self, status_file: str = 'status.json'):
        """
        Saves the account statuses to a JSON file.
        
        Args:
            status_file (str): The path to the status file.
        """
        logger.info(f"Saving account statuses to {status_file}")
        status_data = {}
        for account in self.accounts:
            private_key = account.settings.get('private_key')
            if private_key:
                status_data[private_key] = account.to_dict()
        # Convert int64 to int for JSON serialization
        def convert_int64(obj):
            if isinstance(obj, np.integer):
                return int(obj)
            raise TypeError(f'Object of type {obj.__class__.__name__} is not JSON serializable')
        with open(status_file, 'w') as f:
            json.dump(status_data, f, indent=4, default=convert_int64)
        logger.info("Account statuses saved successfully.")